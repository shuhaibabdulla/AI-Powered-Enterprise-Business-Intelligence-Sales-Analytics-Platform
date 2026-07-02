# ================================================
# RETAILX CORP — AUTOMATED EXCEL REPORT GENERATOR
# Run: python src/reports/excel_report.py
# Output: outputs/excel/retailx_report.xlsx
# ================================================

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
import os
from datetime import datetime

# ---- PATHS ----
EXPORTS_PATH  = 'outputs/exports/'
OUTPUT_PATH   = 'outputs/excel/'
os.makedirs(OUTPUT_PATH, exist_ok=True)

# ---- LOAD DATA ----
print("📂 Loading data...")
fact       = pd.read_csv(EXPORTS_PATH + 'fact_table.csv')
customers  = pd.read_csv(EXPORTS_PATH + 'customer_kpis.csv')
categories = pd.read_csv(EXPORTS_PATH + 'category_kpis.csv')
monthly    = pd.read_csv(EXPORTS_PATH + 'monthly_revenue.csv')
sellers    = pd.read_csv(EXPORTS_PATH + 'seller_kpis.csv')
forecast   = pd.read_csv(EXPORTS_PATH + 'sales_forecast.csv')
executive  = pd.read_csv(EXPORTS_PATH + 'executive_kpi_summary.csv')

print("✅ Data loaded")

# ---- CREATE WORKBOOK ----
wb = Workbook()

# ---- STYLES ----
# Header style
DARK_BLUE   = "1E3A5F"
LIGHT_BLUE  = "2563EB"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F8F9FA"
DARK_GRAY   = "343A40"
GREEN       = "198754"
RED         = "DC3545"
YELLOW      = "FFC107"

def style_header_cell(cell, text, bg_color=DARK_BLUE, font_color=WHITE, size=12):
    cell.value = text
    cell.font = Font(
        bold=True, color=font_color,
        size=size, name='Calibri'
    )
    cell.fill = PatternFill(
        start_color=bg_color,
        end_color=bg_color,
        fill_type='solid'
    )
    cell.alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )

def style_data_cell(cell, value, bold=False, bg_color=None, 
                    font_color=DARK_GRAY, align='center'):
    cell.value = value
    cell.font = Font(bold=bold, color=font_color, 
                     size=11, name='Calibri')
    cell.alignment = Alignment(
        horizontal=align, vertical='center'
    )
    if bg_color:
        cell.fill = PatternFill(
            start_color=bg_color,
            end_color=bg_color,
            fill_type='solid'
        )

def add_border(cell):
    thin = Side(style='thin', color='DDDDDD')
    cell.border = Border(
        left=thin, right=thin,
        top=thin, bottom=thin
    )

# ================================================
# SHEET 1: EXECUTIVE SUMMARY
# ================================================
print("📊 Creating Executive Summary sheet...")
ws1 = wb.active
ws1.title = "Executive Summary"

# Title
ws1.merge_cells('A1:D1')
title_cell = ws1['A1']
title_cell.value = "RetailX Corp — Executive KPI Summary"
title_cell.font = Font(
    bold=True, size=18,
    color=WHITE, name='Calibri'
)
title_cell.fill = PatternFill(
    start_color=DARK_BLUE,
    end_color=DARK_BLUE,
    fill_type='solid'
)
title_cell.alignment = Alignment(
    horizontal='center', vertical='center'
)
ws1.row_dimensions[1].height = 45

# Date
ws1.merge_cells('A2:D2')
date_cell = ws1['A2']
date_cell.value = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
date_cell.font = Font(
    italic=True, size=10,
    color=DARK_GRAY, name='Calibri'
)
date_cell.alignment = Alignment(horizontal='center')

# Column headers
ws1.row_dimensions[4].height = 30
style_header_cell(ws1['A4'], 'KPI', LIGHT_BLUE)
style_header_cell(ws1['B4'], 'Value', LIGHT_BLUE)
style_header_cell(ws1['C4'], 'Category', LIGHT_BLUE)
style_header_cell(ws1['D4'], 'Status', LIGHT_BLUE)

# KPI data with categories and status
kpi_data = [
    ('Total Revenue',           'R$ 15,843,553.24', 'Revenue',   '✅ Strong'),
    ('Total Orders',            '99,441',            'Volume',    '✅ Strong'),
    ('Average Order Value',     'R$ 160.58',         'Revenue',   '✅ Good'),
    ('Total Unique Customers',  '96,096',            'Customer',  '✅ Strong'),
    ('Revenue per Customer',    'R$ 164.87',         'Customer',  '✅ Good'),
    ('Avg Customer Lifetime V', 'R$ 166.04',         'Customer',  '✅ Good'),
    ('Repeat Purchase Rate',    '3.1%',              'Customer',  '⚠️ Low'),
    ('Avg Delivery Days',       '12.0 days',         'Operations','⚠️ Monitor'),
    ('Late Delivery Rate',      '8.1%',              'Operations','⚠️ Monitor'),
    ('Average Review Score',    '4.09 / 5.00',       'Customer',  '✅ Good'),
    ('Total Sellers',           '3,095',             'Supply',    '✅ Good'),
    ('Total Product Categories','72',                'Product',   '✅ Good'),
    ('YoY Revenue Growth',      '21.0%',             'Revenue',   '✅ Strong'),
]

for i, (kpi, value, category, status) in enumerate(kpi_data):
    row = i + 5
    ws1.row_dimensions[row].height = 25
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE

    style_data_cell(ws1[f'A{row}'], kpi,
                    bold=True, bg_color=bg, align='left')
    style_data_cell(ws1[f'B{row}'], value,
                    bg_color=bg, font_color=LIGHT_BLUE)
    style_data_cell(ws1[f'C{row}'], category, bg_color=bg)
    style_data_cell(ws1[f'D{row}'], status, bg_color=bg)

    for col in ['A','B','C','D']:
        add_border(ws1[f'{col}{row}'])

# Column widths
ws1.column_dimensions['A'].width = 30
ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 15

print("✅ Executive Summary sheet done")

# ================================================
# SHEET 2: MONTHLY REVENUE
# ================================================
print("📊 Creating Monthly Revenue sheet...")
ws2 = wb.create_sheet("Monthly Revenue")

# Title
ws2.merge_cells('A1:D1')
ws2['A1'].value = "Monthly Revenue Analysis"
ws2['A1'].font = Font(bold=True, size=16,
                      color=WHITE, name='Calibri')
ws2['A1'].fill = PatternFill(start_color=DARK_BLUE,
                              end_color=DARK_BLUE,
                              fill_type='solid')
ws2['A1'].alignment = Alignment(horizontal='center',
                                 vertical='center')
ws2.row_dimensions[1].height = 40

# Headers
headers = ['Year', 'Month', 'Revenue (R$)', 'MoM Growth']
for i, h in enumerate(headers):
    cell = ws2.cell(row=3, column=i+1)
    style_header_cell(cell, h, LIGHT_BLUE)

# Sort and add data
monthly_sorted = monthly.sort_values(
    ['order_year', 'order_month']
).reset_index(drop=True)
monthly_sorted['mom_growth'] = (
    monthly_sorted['total_revenue'].pct_change() * 100
)

for i, row in monthly_sorted.iterrows():
    r = i + 4
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    ws2.row_dimensions[r].height = 22

    style_data_cell(ws2.cell(r, 1),
                    int(row['order_year']), bg_color=bg)
    style_data_cell(ws2.cell(r, 2),
                    int(row['order_month']), bg_color=bg)
    style_data_cell(ws2.cell(r, 3),
                    f"R$ {row['total_revenue']:,.2f}",
                    bg_color=bg, font_color=LIGHT_BLUE)

    if pd.notna(row['mom_growth']):
        growth = row['mom_growth']
        color  = GREEN if growth >= 0 else RED
        style_data_cell(
            ws2.cell(r, 4),
            f"{growth:+.1f}%",
            bg_color=bg,
            font_color=color
        )

    for col in range(1, 5):
        add_border(ws2.cell(r, col))

# Column widths
ws2.column_dimensions['A'].width = 10
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 20
ws2.column_dimensions['D'].width = 15

# Add Bar Chart
chart = BarChart()
chart.title = "Monthly Revenue"
chart.y_axis.title = "Revenue (R$)"
chart.x_axis.title = "Month"
chart.height = 15
chart.width  = 25

data_ref = Reference(
    ws2,
    min_col=3,
    min_row=3,
    max_row=3 + len(monthly_sorted)
)
chart.add_data(data_ref, titles_from_data=True)
ws2.add_chart(chart, "F3")

print("✅ Monthly Revenue sheet done")

# ================================================
# SHEET 3: CATEGORY PERFORMANCE
# ================================================
print("📊 Creating Category Performance sheet...")
ws3 = wb.create_sheet("Category Performance")

ws3.merge_cells('A1:E1')
ws3['A1'].value = "Product Category Performance"
ws3['A1'].font = Font(bold=True, size=16,
                      color=WHITE, name='Calibri')
ws3['A1'].fill = PatternFill(start_color=DARK_BLUE,
                              end_color=DARK_BLUE,
                              fill_type='solid')
ws3['A1'].alignment = Alignment(horizontal='center',
                                 vertical='center')
ws3.row_dimensions[1].height = 40

headers = ['Rank', 'Category', 'Revenue (R$)',
           'Orders', 'Avg Price (R$)']
for i, h in enumerate(headers):
    cell = ws3.cell(row=3, column=i+1)
    style_header_cell(cell, h, LIGHT_BLUE)

cat_sorted = categories.sort_values(
    'total_revenue', ascending=False
).reset_index(drop=True)

for i, row in cat_sorted.iterrows():
    r = i + 4
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    ws3.row_dimensions[r].height = 22

    # Rank medal for top 3
    rank = i + 1
    rank_label = (
        '🥇' if rank == 1 else
        '🥈' if rank == 2 else
        '🥉' if rank == 3 else
        str(rank)
    )

    style_data_cell(ws3.cell(r,1), rank_label, bg_color=bg)
    style_data_cell(ws3.cell(r,2),
                    row['product_category_name_english'],
                    bg_color=bg, align='left')
    style_data_cell(ws3.cell(r,3),
                    f"R$ {row['total_revenue']:,.2f}",
                    bg_color=bg, font_color=LIGHT_BLUE)
    style_data_cell(ws3.cell(r,4),
                    f"{int(row['total_orders']):,}",
                    bg_color=bg)
    style_data_cell(ws3.cell(r,5),
                    f"R$ {row['avg_price']:,.2f}",
                    bg_color=bg)

    for col in range(1, 6):
        add_border(ws3.cell(r, col))

ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 30
ws3.column_dimensions['C'].width = 18
ws3.column_dimensions['D'].width = 12
ws3.column_dimensions['E'].width = 16

print("✅ Category Performance sheet done")

# ================================================
# SHEET 4: SALES FORECAST
# ================================================
print("📊 Creating Forecast sheet...")
ws4 = wb.create_sheet("Sales Forecast")

ws4.merge_cells('A1:C1')
ws4['A1'].value = "3-Month Revenue Forecast"
ws4['A1'].font = Font(bold=True, size=16,
                      color=WHITE, name='Calibri')
ws4['A1'].fill = PatternFill(start_color=DARK_BLUE,
                              end_color=DARK_BLUE,
                              fill_type='solid')
ws4['A1'].alignment = Alignment(horizontal='center',
                                 vertical='center')
ws4.row_dimensions[1].height = 40

headers = ['Month', 'Predicted Revenue (R$)', 'Type']
for i, h in enumerate(headers):
    cell = ws4.cell(row=3, column=i+1)
    style_header_cell(cell, h, LIGHT_BLUE)

for i, row in forecast.iterrows():
    r = i + 4
    ws4.row_dimensions[r].height = 25
    style_data_cell(ws4.cell(r,1), row['month'])
    style_data_cell(
        ws4.cell(r,2),
        f"R$ {row['predicted_revenue']:,.2f}",
        font_color=LIGHT_BLUE, bold=True
    )
    style_data_cell(ws4.cell(r,3), '🔮 Forecast',
                    font_color=RED)

    for col in range(1, 4):
        add_border(ws4.cell(r, col))

ws4.column_dimensions['A'].width = 15
ws4.column_dimensions['B'].width = 25
ws4.column_dimensions['C'].width = 15

print("✅ Forecast sheet done")

# ================================================
# SAVE WORKBOOK
# ================================================
output_file = OUTPUT_PATH + 'retailx_report.xlsx'
wb.save(output_file)

print()
print("=" * 50)
print("✅ EXCEL REPORT GENERATED SUCCESSFULLY!")
print(f"   File: {output_file}")
print(f"   Sheets: 4")
print("   1. Executive Summary")
print("   2. Monthly Revenue")
print("   3. Category Performance")
print("   4. Sales Forecast")
print("=" * 50)